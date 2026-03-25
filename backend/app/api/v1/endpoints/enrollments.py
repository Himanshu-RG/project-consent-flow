"""
Project enrollment endpoints for managing user-project assignments.
Allows admins to enroll users in projects for ML face matching.
"""

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List
from uuid import UUID
from app.db.database import get_db
from app.models.models import Project, User, project_enrollments
from app.schemas.schemas import UserResponse, MessageResponse


router = APIRouter()


@router.post("/projects/{project_id}/enroll/{user_id}", response_model=MessageResponse)
async def enroll_user_in_project(
    project_id: UUID,
    user_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Enroll a user in a project.
    This allows the user's face to be matched against detected faces in project images.
    """
    # Check if project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    # Check if user exists
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Check if already enrolled
    if user in project.enrolled_users:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User is already enrolled in this project"
        )
    
    # Enroll user
    project.enrolled_users.append(user)
    db.commit()
    
    return {"message": f"User {user.email} enrolled in project {project.name}"}


@router.delete("/projects/{project_id}/enroll/{user_id}", response_model=MessageResponse)
async def unenroll_user_from_project(
    project_id: UUID,
    user_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Remove a user from a project.
    """
    # Check if project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    # Check if user exists
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Check if enrolled
    if user not in project.enrolled_users:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User is not enrolled in this project"
        )
    
    # Unenroll user
    project.enrolled_users.remove(user)
    db.commit()
    
    return {"message": f"User {user.email} removed from project {project.name}"}


@router.get("/projects/{project_id}/enrolled-users", response_model=List[UserResponse])
async def get_enrolled_users(
    project_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Get all users enrolled in a project.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    return project.enrolled_users


@router.get("/projects/{project_id}/enrollment-status", response_model=List)
async def get_enrollment_status(
    project_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Get enrollment and consent status for all users enrolled in a project.
    Shows which users have been detected, their consent status, and match confidence.
    """
    from app.models.models import Person
    from app.schemas.schemas import EnrollmentStatusResponse
    
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    enrollment_statuses = []
    
    for user in project.enrolled_users:
        # Check if user has been detected in project (has Person record)
        person = db.query(Person).filter(
            Person.project_id == project_id,
            Person.user_id == user.id
        ).first()
        
        is_detected = person is not None
        
        # Determine consent status
        if not is_detected:
            consent_status = "pending"  # Not yet detected in images
        elif person and person.consent_status == "granted":
            consent_status = "matching"  # Person has consent (either from user or direct upload)
        elif user.consent_pdf_url:
            consent_status = "matching"  # User has global consent PDF
        else:
            consent_status = "not_matching"  # Detected but no consent
        
        status_obj = EnrollmentStatusResponse(
            user_id=user.id,
            user_name=user.full_name or "Unknown",
            user_email=user.email,
            pid=user.pid,
            is_detected=is_detected,
            consent_status=consent_status,
            match_confidence=person.match_confidence if person else None,
            person_id=person.id if person else None,
            has_identity_image=bool(user.identity_image_url),
            has_consent_pdf=bool(user.consent_pdf_url)
        )
        
        enrollment_statuses.append(status_obj)
    
    return enrollment_statuses


@router.get("/projects/{project_id}/enrollment-status/export")
async def export_enrollment_status(
    project_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Export enrollment and consent status to Excel file.
    Returns a downloadable Excel file with all enrollment data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from app.models.models import Person
    
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Consent Status"
    
    # Define headers
    headers = ["User Name", "Email", "PID", "Detected", "Consent Status", "Match Confidence (%)", "Has Identity Image", "Has Consent PDF", "Consent PDF Location"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for user in project.enrolled_users:
        person = db.query(Person).filter(
            Person.project_id == project_id,
            Person.user_id == user.id
        ).first()
        
        is_detected = person is not None
        
        # Determine consent status
        if not is_detected:
            consent_status = "Pending Detection"
        elif person and person.consent_status == "granted":
            consent_status = "Matching"
        elif user.consent_pdf_url:
            consent_status = "Matching"
        else:
            consent_status = "Not Matching"
        
        # Check for consent PDF - either from user or person's consent form
        has_consent = False
        consent_location = "-"
        
        if user.consent_pdf_url:
            has_consent = True
            consent_location = user.consent_pdf_path or user.consent_pdf_url
        elif person and person.consent_status == "granted":
            # Check if person has a consent form
            from app.models.models import ConsentForm
            consent_form = db.query(ConsentForm).filter(
                ConsentForm.person_id == person.id
            ).first()
            if consent_form:
                has_consent = True
                consent_location = consent_form.file_path or consent_form.file_url
        
        row = [
            user.full_name or "Unknown",
            user.email,
            user.pid or "-",
            "Yes" if is_detected else "No",
            consent_status,
            person.match_confidence if person and person.match_confidence else "-",
            "Yes" if user.identity_image_url else "No",
            "Yes" if has_consent else "No",
            consent_location
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Return as downloadable file
    filename = f"{project.name.replace(' ', '_')}_consent_status.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



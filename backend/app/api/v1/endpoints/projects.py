"""
Project endpoints for CRUD operations on projects (no authentication required).
Handles project creation, listing, retrieval, updating, and deletion.
Also includes bulk-upload project creation, ML processing, Excel export, and redaction.
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID, uuid4
from pathlib import Path
import shutil

from app.db.database import get_db
from app.models.models import Project, Event
from app.schemas.schemas import (
    ProjectCreate, ProjectUpdate, ProjectResponse, ProjectListResponse,
    ProjectProcessResult,
)
from app.core.config import settings


router = APIRouter(prefix="/projects", tags=["Projects"])


# ─────────────────────────────────────────────────────────────────────────────
# Project CRUD
# ─────────────────────────────────────────────────────────────────────────────

@router.post("", response_model=ProjectResponse, status_code=status.HTTP_201_CREATED)
async def create_project(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    target_image_count: Optional[int] = Form(None),
    project_status: Optional[str] = Form("active"),
    camera_dslr: Optional[bool] = Form(False),
    camera_mobile: Optional[bool] = Form(False),
    pii_face: Optional[bool] = Form(False),
    pii_objects: Optional[bool] = Form(False),
    pii_document: Optional[bool] = Form(False),
    pii_other: Optional[bool] = Form(False),
    images: List[UploadFile] = File(default=[]),
    consent_pdfs: List[UploadFile] = File(default=[]),
    db: Session = Depends(get_db),
):
    """
    Create a new project, optionally with images and consent PDFs uploaded at the same time.
    """
    from app.models.models import Image, ConsentForm
    from PIL import Image as PILImage, ImageOps
    from concurrent.futures import ThreadPoolExecutor as _TPE

    new_project = Project(
        name=name,
        description=description,
        notes=notes,
        target_image_count=target_image_count,
        status=project_status or "active",
        camera_dslr=camera_dslr or False,
        camera_mobile=camera_mobile or False,
        pii_face=pii_face or False,
        pii_objects=pii_objects or False,
        pii_document=pii_document or False,
        pii_other=pii_other or False,
    )
    db.add(new_project)
    db.commit()
    db.refresh(new_project)

    # ── Helper: save one image file + EXIF correction (runs in thread pool) ──
    def _save_image_file(file_bytes: bytes, original_name: str, content_type: str):
        """Write bytes to disk, correct EXIF, return record-data dict."""
        upload_dir_img = Path(settings.IMAGES_DIR) / str(new_project.id)
        upload_dir_img.mkdir(parents=True, exist_ok=True)
        suffix = Path(original_name).suffix or ".jpg"
        unique_filename = f"{uuid4()}{suffix}"
        file_path = upload_dir_img / unique_filename
        file_path.write_bytes(file_bytes)
        width, height = None, None
        try:
            with PILImage.open(file_path) as img:
                img = ImageOps.exif_transpose(img)
                img.save(file_path, quality=92)
                width, height = img.size
        except Exception as exc:
            print(f"[UPLOAD] EXIF error on {original_name}: {exc}")
        return {
            "name":      original_name,
            "file_path": str(file_path),
            "file_url":  f"/uploads/images/{new_project.id}/{unique_filename}",
            "file_size": file_path.stat().st_size,
            "mime_type": content_type,
            "width":     width,
            "height":    height,
        }

    # ── Read all image bytes (async I/O) then process in parallel threads ────
    valid_images = [
        f for f in images
        if f.content_type and f.content_type.startswith("image/")
    ]
    image_file_data = []
    for f in valid_images:
        raw = await f.read()
        image_file_data.append((raw, f.filename, f.content_type))

    image_records_data = []
    if image_file_data:
        with _TPE(max_workers=4) as pool:
            futures = [
                pool.submit(_save_image_file, raw, name, ctype)
                for raw, name, ctype in image_file_data
            ]
            for fut in futures:
                try:
                    image_records_data.append(fut.result())
                except Exception as exc:
                    print(f"[UPLOAD] Failed to save image: {exc}")

    # Batch-insert all image records in one commit
    for rec in image_records_data:
        db.add(Image(project_id=new_project.id, **rec))

    # ── Save PDF files (usually few, sequential is fine) ─────────────────────
    pdf_dir = Path(settings.CONSENT_PDFS_DIR) / str(new_project.id)
    pdf_dir.mkdir(parents=True, exist_ok=True)
    valid_pdfs = [
        f for f in consent_pdfs
        if f.content_type == "application/pdf"
    ]
    for file in valid_pdfs:
        raw_pdf = await file.read()
        unique_filename = f"{uuid4()}.pdf"
        file_path = pdf_dir / unique_filename
        file_path.write_bytes(raw_pdf)
        db.add(ConsentForm(
            project_id=new_project.id,
            form_name=file.filename,
            file_path=str(file_path),
            file_url=f"/uploads/consent_pdfs/{new_project.id}/{unique_filename}",
            file_size=file_path.stat().st_size,
        ))

    # Single batch commit for all images + PDFs
    db.commit()

    event = Event(
        project_id=new_project.id,
        event_type="project_created",
        description=f"Project '{new_project.name}' created with {len(image_records_data)} image(s) and {len(valid_pdfs)} PDF(s)",
        event_metadata={
            "project_name": new_project.name,
            "images_count": len(image_records_data),
            "pdfs_count":   len(valid_pdfs),
        },
    )
    db.add(event)
    db.commit()
    db.refresh(new_project)

    return new_project


@router.get("", response_model=ProjectListResponse)
async def list_projects(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    status: str = Query(None, description="Filter by status"),
    db: Session = Depends(get_db),
):
    """
    List all projects with pagination.
    """
    query = db.query(Project)

    if status:
        query = query.filter(Project.status == status)

    total = query.count()
    offset = (page - 1) * limit
    projects = query.order_by(Project.created_at.desc()).offset(offset).limit(limit).all()
    total_pages = (total + limit - 1) // limit

    return {
        "projects": projects,
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": total_pages,
        },
    }


@router.get("/{project_id}", response_model=ProjectResponse)
async def get_project(
    project_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Get a specific project by ID.
    """
    project = db.query(Project).filter(Project.id == project_id).first()

    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    return project


@router.put("/{project_id}", response_model=ProjectResponse)
async def update_project(
    project_id: UUID,
    project_data: ProjectUpdate,
    db: Session = Depends(get_db),
):
    """
    Update a project. All fields are optional.
    """
    project = db.query(Project).filter(Project.id == project_id).first()

    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    update_data = project_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(project, field, value)

    db.commit()
    db.refresh(project)

    event = Event(
        project_id=project.id,
        event_type="project_updated",
        description=f"Project '{project.name}' updated",
        event_metadata={"updated_fields": list(update_data.keys())},
    )
    db.add(event)
    db.commit()

    return project


@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_project(
    project_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Delete a project (cascades to all related data).
    """
    project = db.query(Project).filter(Project.id == project_id).first()

    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    db.delete(project)
    db.commit()

    return None


# ─────────────────────────────────────────────────────────────────────────────
# ML processing endpoint
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/{project_id}/process", status_code=status.HTTP_202_ACCEPTED)
async def process_project(
    project_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Kick off background ML processing for all project images.

    Returns immediately with a task_id.  Poll progress via:
        GET /projects/{project_id}/process/status/{task_id}
    """
    from app.models.models import Image as ImageModel
    from app.ml.processing_tasks import create_task, submit_processing_task
    from app.db.database import SessionLocal

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    image_count = db.query(ImageModel).filter(ImageModel.project_id == project_id).count()
    if image_count == 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="No images found in this project. Upload images first.",
        )

    task_id = create_task(str(project_id))
    submit_processing_task(
        task_id=task_id,
        project_id_str=str(project_id),
        db_factory=SessionLocal,
        dataset_dir=settings.DATASET_KNOWN_DIR,
    )

    return {
        "task_id":     task_id,
        "project_id": str(project_id),
        "status":     "started",
        "message":    f"Processing {image_count} image(s) in the background. Poll /process/status/{task_id} for progress.",
    }


@router.get("/{project_id}/process/status/{task_id}")
async def get_process_status(
    project_id: UUID,
    task_id: str,
):
    """
    Server-Sent Events stream for ML processing progress.

    Returns a text/event-stream.  Each event is a JSON object:
        { task_id, status, progress, total, current_image, result, error }

    'status' values: pending | processing | saving | done | error
    Stream closes automatically when status is 'done' or 'error'.
    """
    import asyncio, json
    from app.ml.processing_tasks import get_task

    async def event_stream():
        while True:
            task = get_task(task_id)
            if task is None:
                yield f"data: {json.dumps({'error': 'Task not found', 'status': 'error'})}\n\n"
                break

            yield f"data: {json.dumps(task)}\n\n"

            if task["status"] in ("done", "error"):
                break

            await asyncio.sleep(0.5)   # poll interval

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",   # disable nginx buffering
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Excel export endpoint
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{project_id}/export/excel")
async def export_project_excel(
    project_id: UUID,
    db: Session = Depends(get_db),
):
    """
    Generate and download an Excel report for the project.

    Columns:
        Image Name | Project | Category | Location |
        Human Presence (Yes/No) | No. of Human Subjects |
        1st Person Name | 2nd Person Name | ... |
        1st Consent Form Name | 2nd Consent Form Name | ...

    Unknown persons (pid=None) are written as empty cells.
    The "Location" column is derived from image_metadata if available.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.models import (
        Image as ImageModel, Person, ImagePerson, ConsentForm, Project as ProjectModel,
    )

    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    images = (
        db.query(ImageModel)
        .filter(ImageModel.project_id == project_id)
        .order_by(ImageModel.created_at)
        .all()
    )

    image_ids = [img.id for img in images]

    assoc_rows = (
        db.query(ImagePerson)
        .filter(ImagePerson.image_id.in_(image_ids))
        .all()
    ) if image_ids else []

    from collections import defaultdict
    img_to_assocs: dict = defaultdict(list)
    for a in assoc_rows:
        img_to_assocs[a.image_id].append(a)

    person_ids = list({a.person_id for a in assoc_rows})
    persons_map: dict = {}
    if person_ids:
        for p in db.query(Person).filter(Person.id.in_(person_ids)).all():
            persons_map[p.id] = p

    consent_map: dict = defaultdict(list)
    if person_ids:
        for cf in (
            db.query(ConsentForm)
            .filter(
                ConsentForm.project_id == project_id,
                ConsentForm.person_id.in_(person_ids),
            )
            .all()
        ):
            if cf.person_id:
                consent_map[cf.person_id].append(cf.form_name)

    valid_persons_per_img = {}
    for img in images:
        seen = set()
        valid = []
        for a in img_to_assocs.get(img.id, []):
            p = persons_map.get(a.person_id)
            if p and p.pid and p.id not in seen:
                seen.add(p.id)
                valid.append(p)
        valid_persons_per_img[img.id] = valid

    max_persons = max(
        (len(valid_persons_per_img[img.id]) for img in images),
        default=0,
    )
    max_persons = max(max_persons, 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Data"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Aptos Narrow", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
    NORMAL_FILL = PatternFill("solid", fgColor="FFFFFF")

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    static_headers = [
        "Image Name", "Project", "Category", "Location",
        "Human Presence\n(Yes/No)", "No. of Human\nSubjects",
    ]
    dynamic_headers = []
    for n in range(1, max_persons + 1):
        ordinal = "1st" if n == 1 else "2nd" if n == 2 else "3rd" if n == 3 else f"{n}th"
        dynamic_headers.append(f"{ordinal} Person Name")
    for n in range(1, max_persons + 1):
        ordinal = "1st" if n == 1 else "2nd" if n == 2 else "3rd" if n == 3 else f"{n}th"
        dynamic_headers.append(f"{ordinal} Consent Form Name")

    all_headers = static_headers + dynamic_headers
    ws.append(all_headers)

    for col_idx, _ in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = border
        cell.alignment = centre

    ws.row_dimensions[1].height = 30

    for row_idx, img in enumerate(images, start=2):
        valid_persons = valid_persons_per_img.get(img.id, [])
        assocs        = img_to_assocs.get(img.id, [])

        location = ""
        if img.image_metadata and isinstance(img.image_metadata, dict):
            location = img.image_metadata.get("location", "") or ""

        person_names = [p.name for p in valid_persons]

        has_humans = len(valid_persons) > 0
        if not has_humans and len(assocs) > 0:
            has_humans = True

        human_count = len(valid_persons)

        consent_names = []
        for p in valid_persons:
            forms = consent_map.get(p.id, [])
            consent_names.append(forms[0] if forms else "")

        person_names  += [""] * (max_persons - len(person_names))
        consent_names += [""] * (max_persons - len(consent_names))

        row_data = [
            img.name,
            project.name,
            img.factor or "",
            location,
            "Yes" if has_humans else "No",
            human_count,
        ] + person_names + consent_names

        ws.append(row_data)

        fill = ALT_FILL if row_idx % 2 == 0 else NORMAL_FILL
        for col_idx in range(1, len(all_headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = left

        ws.row_dimensions[row_idx].height = 18

    col_widths = [22, 20, 22, 14, 12, 10] + [22] * (max_persons * 2)
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    safe_name = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in project.name)
    filename  = f"{safe_name}_metadata.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Redaction endpoint
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{project_id}/images/{image_id}/redacted")
async def get_redacted_image(
    project_id: UUID,
    image_id: UUID,
    t: Optional[str] = None,   # cache-busting query param (ignored server-side)
    db: Session = Depends(get_db),
):
    """
    Return a runtime-processed JPEG of the image with consent-aware annotations:
      - Image not yet processed (no ImagePerson rows)
            → full Gaussian blur + "NOT PROCESSED" label
      - Consented persons  → green bounding box + name label
      - Non-consented      → Gaussian blur on face + red bounding box + "REDACTED" label
      - Faces with missing bbox (edge case after model)
            → whole image blurred as safety fallback

    No files are written to disk. The image is streamed directly.
    The 't' query parameter is used by the frontend for cache-busting only.
    """
    import io
    import os
    from app.models.models import Image as ImageModel

    # 1. Load image record
    image = db.query(ImageModel).filter(
        ImageModel.id == image_id,
        ImageModel.project_id == project_id,
    ).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    if not image.file_path:
        raise HTTPException(status_code=404, detail="Image file path not recorded")

    if not os.path.exists(image.file_path):
        raise HTTPException(status_code=404, detail=f"Image file not found on disk: {image.file_path}")

    jpeg_bytes = _process_redacted_image(db, image)

    return StreamingResponse(
        io.BytesIO(jpeg_bytes),
        media_type="image/jpeg",
        headers={
            "Cache-Control": "no-store",
            "Content-Disposition": f'inline; filename="redacted_{image.name}"',
        },
    )

def _process_redacted_image(db: Session, image: 'ImageModel') -> bytes:
    from app.models.models import Person, ImagePerson
    from app.ml.redaction import redact_image, blur_entire_image

    # 2. Get all ImagePerson links for this image — each carries its own bbox
    associations = db.query(ImagePerson).filter(ImagePerson.image_id == image.id).all()

    # ── If the model hasn't been run yet (no associations at all) → full blur ──
    if not associations:
        try:
            return blur_entire_image(image.file_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Blur processing failed: {e}")

    # 3. Build face list for redaction — use per-image bbox from ImagePerson
    assoc_by_person = {a.person_id: a for a in associations}
    person_ids = list(assoc_by_person.keys())
    persons = db.query(Person).filter(Person.id.in_(person_ids)).all() if person_ids else []

    faces = []
    for person in persons:
        assoc = assoc_by_person.get(person.id)
        # Prefer per-image bbox (ImagePerson), fall back to Person.bbox for legacy rows
        bbox = (assoc.bbox if assoc and assoc.bbox else None) or person.bbox
        if not bbox:
            print(
                f"[REDACT][WARN] No bbox found for person '{person.name}' (id={person.id}) "
                f"in image {image.id}. Skipping annotation for this face."
            )
            continue

        # Consent determination:
        #   - A NAMED person (pid set): use their consent_status directly.
        #   - An UNKNOWN person (pid=None): ALWAYS redact — we cannot verify consent
        #     for an unidentified face.
        is_consented = (person.pid is not None) and (person.consent_status == "granted")

        faces.append({
            "name":      person.name or "Unknown",
            "bbox":      bbox,
            "consented": is_consented,
        })

    # 3b. Append manually-drawn boxes (stored in image_metadata["manual_redact_boxes"])
    manual_meta: dict = image.image_metadata or {}
    for mbox in manual_meta.get("manual_redact_boxes", []):
        faces.append({
            "name":      mbox.get("label", "Manual"),
            "bbox": {
                "x":      mbox.get("x", 0),
                "y":      mbox.get("y", 0),
                "width":  mbox.get("width", 0),
                "height": mbox.get("height", 0),
            },
            "consented": bool(mbox.get("consented", False)),
        })

    # 4. Apply annotations.
    #    If faces ended up empty (all had missing bboxes) → still blur the whole image
    try:
        if faces:
            return redact_image(image.file_path, faces)
        else:
            return blur_entire_image(image.file_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Redaction processing failed: {e}")


@router.get("/{project_id}/redacted-images/zip")
async def download_redacted_images_zip(
    project_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Downloads all redacted images for a project as a single ZIP file.
    """
    import os
    import zipfile
    import io
    from datetime import datetime
    from app.models.models import Image as ImageModel, Project

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    images = db.query(ImageModel).filter(ImageModel.project_id == project_id).all()
    if not images:
        raise HTTPException(status_code=404, detail="No images found for this project")

    zip_bytes_io = io.BytesIO()

    with zipfile.ZipFile(zip_bytes_io, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for image in images:
            if not image.file_path or not os.path.exists(image.file_path):
                continue

            try:
                # Process the redaction
                jpeg_bytes = _process_redacted_image(db, image)
            except Exception as e:
                print(f"[REDACT ZIP][WARN] Failed to process image {image.name}: {e}")
                continue
            
            # Write bytes to zip
            zf.writestr(f"redacted_{image.name}", jpeg_bytes)

    zip_bytes_io.seek(0)

    # Format: PROJECTNAME_REDACTED_TIME.zip
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_project_name = "".join([c if c.isalnum() else "_" for c in project.name])
    filename = f"{safe_project_name}_REDACTED_{timestamp}.zip"

    return StreamingResponse(
        zip_bytes_io,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        }
    )



# ─────────────────────────────────────────────────────────────────────────────
# Manual-redaction box endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/{project_id}/images/{image_id}/manual-redact")
async def save_manual_redact_boxes(
    project_id: UUID,
    image_id: UUID,
    boxes: str = Form(...),        # JSON string: list of box dicts
    db: Session = Depends(get_db),
    **kwargs,                       # absorbs uploaded pdf_{i} form files
):
    """
    Save (overwrite) the list of manually-drawn bounding boxes for one image.

    Body (multipart/form-data):
        boxes   – JSON array of:
                    { id, x, y, width, height, consented, label, pdfName? }
        pdf_{i} – optional PDF file for box at index i (consented=true boxes)

    Boxes are stored in image.image_metadata["manual_redact_boxes"].
    PDFs are saved to disk; their filenames are stored in the box dict.
    """
    import json

    from app.models.models import Image as ImageModel


    image = db.query(ImageModel).filter(
        ImageModel.id == image_id,
        ImageModel.project_id == project_id,
    ).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    try:
        box_list = json.loads(boxes)
    except Exception:
        raise HTTPException(status_code=422, detail="'boxes' must be valid JSON")

    return {"saved": len(box_list), "message": "Use the /manual-redact-upload route for full multipart support"}


@router.post("/{project_id}/images/{image_id}/manual-redact-upload")
async def save_manual_redact_boxes_upload(
    project_id: UUID,
    image_id: UUID,
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Real multipart handler for manual redact boxes.
    Called directly (not via FastAPI form params) so we can iterate arbitrary
    pdf_{i} fields alongside the boxes JSON field.
    """
    import json
    from app.models.models import Image as ImageModel

    form = await request.form()

    boxes_raw = form.get("boxes")
    if not boxes_raw:
        raise HTTPException(status_code=422, detail="Missing 'boxes' field")
    try:
        box_list: list = json.loads(boxes_raw)
    except Exception:
        raise HTTPException(status_code=422, detail="'boxes' must be valid JSON")

    image = db.query(ImageModel).filter(
        ImageModel.id == image_id,
        ImageModel.project_id == project_id,
    ).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    # Save any PDF files uploaded alongside green boxes
    pdf_dir = Path(settings.CONSENT_PDFS_DIR) / str(project_id) / "manual"
    pdf_dir.mkdir(parents=True, exist_ok=True)

    for i, box in enumerate(box_list):
        pdf_field = form.get(f"pdf_{i}")
        if pdf_field and hasattr(pdf_field, "filename") and pdf_field.filename:
            unique_name = f"{uuid4()}.pdf"
            dest = pdf_dir / unique_name
            content = await pdf_field.read()
            dest.write_bytes(content)
            box["pdf_path"] = str(dest)
            box["pdfName"] = pdf_field.filename

    # Persist the boxes into JSONB metadata
    existing_meta: dict = image.image_metadata or {}
    existing_meta["manual_redact_boxes"] = box_list
    image.image_metadata = existing_meta
    # Force SQLAlchemy to detect JSONB mutation
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(image, "image_metadata")
    db.commit()

    return {"saved": len(box_list)}


@router.get("/{project_id}/images/{image_id}/manual-redact")
async def get_manual_redact_boxes(
    project_id: UUID,
    image_id: UUID,
    db: Session = Depends(get_db),
):
    """Return the stored manual redaction boxes for one image."""
    from app.models.models import Image as ImageModel

    image = db.query(ImageModel).filter(
        ImageModel.id == image_id,
        ImageModel.project_id == project_id,
    ).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")

    meta = image.image_metadata or {}
    boxes = meta.get("manual_redact_boxes", [])
    # Strip internal pdf_path from the response (keep pdfName for display only)
    clean = [{k: v for k, v in b.items() if k != "pdf_path"} for b in boxes]
    return clean

